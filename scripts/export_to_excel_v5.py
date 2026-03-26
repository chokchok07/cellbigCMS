import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

filepath = r"c:\Users\user\Documents\VSCode\CellbigCMS\CellbigCMS\api_specification.xlsx"

headers = ["카테고리", "메서드", "API 경로", "설명", "요청 파라미터 (Body/Query)", "응답 데이터 (JSON 예시)"]

client_data = [
    ["Device/Client", "POST", "/devices/register", "디바이스 최초 등록 요청", 
     "Body:\n- macAddress\n- systemInfo\n- clientVersion", 
     "{\n  \"success\": true,\n  \"data\": {\n    \"deviceId\": \"dev_123\",\n    \"status\": \"PENDING_APPROVAL\",\n    \"accessToken\": \"eyJ...\"\n  }\n}"],
    ["Device/Client", "PUT", "/devices/{deviceId}/info", "디바이스 정보 정기 갱신 (월/분기별)", 
     "Body:\n- systemInfo(cpu, ram, disk 등)\n- clientVersion\n- networkInfo\n- reportPeriod", 
     "{\n  \"success\": true,\n  \"data\": {\n    \"updatedAt\": \"2026-03-11T10:00:00Z\",\n    \"message\": \"갱신 완료\"\n  }\n}"],
    ["Device/Client", "POST", "/devices/certify", "콘텐츠 실행 인증 (상시/One-time)", 
     "Body:\n- deviceId\n- contentId\n- macAddress\n- authType", 
     "{\n  \"success\": true,\n  \"data\": {\n    \"authorized\": true,\n    \"sessionId\": \"sess_abc123\",\n    \"license\": {\n      \"validUntil\": \"2025-12-31T23:59:59Z\",\n      \"type\": \"SUBSCRIPTION\"\n    }\n  }\n}"],
    ["Device/Client", "GET", "/devices/{deviceId}/packages/check-update", "업데이트 확인", 
     "Query:\n- currentPackageVersion", 
     "{\n  \"success\": true,\n  \"data\": {\n    \"hasUpdate\": true,\n    \"targetPackage\": {...},\n    \"contents\": [...]\n  }\n}"],
    ["Device/Client", "GET", "/contents/{contentId}/versions/{version}/download", "콘텐츠 다운로드 URL 요청", "-", 
     "{\n  \"success\": true,\n  \"data\": {\n    \"downloadUrl\": \"https://...\",\n    \"expiresAt\": \"...\",\n    \"fileSize\": 104857600,\n    \"checksum\": \"...\"\n  }\n}"]
]

admin_data = [
    ["Admin/Auth", "POST", "/auth/login", "관리자 로그인", "Body:\n- username\n- password", "{\n  \"success\": true,\n  \"data\": {\n    \"accessToken\": \"eyJ...\",\n    \"refreshToken\": \"...\"\n  }\n}"],
    ["Admin/Auth", "POST", "/auth/register", "운영자/관리자 회원가입", "Body:\n- username\n- password\n- email\n- fullName", "{\n  \"success\": true,\n  \"data\": {\n    \"userId\": \"user_123\",\n    \"status\": \"PENDING_APPROVAL\",\n    \"message\": \"관리자 승인 대기\"\n  }\n}"],
    ["Admin/Auth", "GET", "/auth/me", "내 정보 및 권한 조회", "-", "{\n  \"success\": true,\n  \"data\": {\n    \"userId\": \"u1\",\n    \"username\": \"admin\",\n    \"role\": \"super_admin\",\n    \"scopeIds\": [],\n    \"permissions\": [\"device:read\"]\n  }\n}"],
    ["Admin/Auth", "PUT", "/auth/me", "내 정보 수정", "Body:\n- currentPassword\n- newPassword\n- fullName", "{\n  \"success\": true,\n  \"data\": {}\n}"],
    
    ["Admin/Users", "GET", "/users", "사용자 목록 조회", "-", "{\n  \"success\": true,\n  \"data\": [\n    { \"userId\": \"u1\", \"username\": \"m1\", \"role\": \"local_manager\", ... }\n  ]\n}"],
    ["Admin/Users", "POST", "/users", "새 운영자 계정 생성", "Body:\n- username\n- password\n- fullName\n- role\n- scopeIds", "{\n  \"success\": true,\n  \"data\": { \"userId\": \"u2\" }\n}"],
    ["Admin/Users", "PUT", "/users/{userId}", "운영자 정보/권한 수정", "Body:\n- role\n- status\n- scopeIds", "{\n  \"success\": true,\n  \"data\": {}\n}"],
    
    ["Admin/Products", "GET", "/products", "제품 목록 조회", "-", "{\n  \"success\": true,\n  \"data\": [ { \"productId\": \"p1\", \"name\": \"...\" } ]\n}"],
    ["Admin/Products", "POST", "/products", "제품 생성", "Body:\n- name\n- description (제품 메타데이터)", "{\n  \"success\": true,\n  \"data\": { \"productId\": \"p2\" }\n}"],
    
    ["Admin/Packages", "GET", "/products/{productId}/packages", "제품 하위 패키지 목록", "-", "{\n  \"success\": true,\n  \"data\": [ { \"packageId\": \"pkg1\", \"name\": \"...\" } ]\n}"],
    ["Admin/Packages", "PUT", "/packages/{packageId}", "패키지 정보 수정 / 콘텐츠 매핑", "Body:\n- name\n- contentIds", "{\n  \"success\": true,\n  \"data\": {}\n}"],
    
    ["Admin/Contents", "POST", "/contents", "콘텐츠 메타데이터 생성", "Body:\n- title\n- type\n- description", "{\n  \"success\": true,\n  \"data\": { \"contentId\": \"c1\" }\n}"],
    ["Admin/Contents", "POST", "/contents/{contentId}/versions", "새 버전 업로드 (파일 포함)", "Multipart Form:\n- file\n- versionMajor/Minor/Patch\n- releaseNotes", "{\n  \"success\": true,\n  \"data\": {\n    \"version\": \"1.1.0\",\n    \"checksum\": \"...\"\n  }\n}"],
    
    ["Admin/LocalArea", "GET", "/local-areas", "지역 목록 (트리 구조)", "-", "{\n  \"success\": true,\n  \"data\": [ { \"localAreaId\": \"loc_1\", \"name\": \"서울\", \"children\": [...] } ]\n}"],
    ["Admin/Store", "POST", "/stores", "스토어 생성", "Body:\n- localAreaId\n- storeName", "{\n  \"success\": true,\n  \"data\": { \"storeId\": \"store_1\" }\n}"],
    
    ["Admin/Devices", "GET", "/admin/devices", "전체 디바이스 목록 (검색/필터)", "Query:\n- status\n- storeId\n- page\n- size", "{\n  \"success\": true,\n  \"data\": { \"content\": [...], \"totalElements\": 100 }\n}"],
    ["Admin/Devices", "PUT", "/admin/devices/{deviceId}/approve", "디바이스 승인 및 스토어 할당", "Body:\n- storeId\n- packageId\n- status", "{\n  \"success\": true,\n  \"data\": {}\n}"],
    
    ["Admin/License", "POST", "/licenses", "라이선스(시리얼키) 발급", "Body:\n- productId\n- type\n- durationDays\n- count", "{\n  \"success\": true,\n  \"data\": [\n    \"License-ABCD-1234\",\n    \"License-EFGH-5678\"\n  ]\n}"]
]

logging_data = [
    ["Logging", "POST", "/logs/access", "콘텐츠 실행/종료 로그 전송", "Body:\n- deviceId\n- sessionId\n- contentId\n- action\n- timestamp\n- meta", "{\n  \"success\": true,\n  \"data\": {}\n}"]
]

wb = Workbook()
wb.remove(wb.active) # Remove default sheet

def create_sheet(title, data):
    ws = wb.create_sheet(title=title)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    alignment = Alignment(wrap_text=True, vertical="top")

    ws.append(headers)
    for row_data in data:
        ws.append(row_data)

    for i, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.alignment = alignment
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill

    col_widths = {'A': 15, 'B': 10, 'C': 45, 'D': 35, 'E': 40, 'F': 50}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

create_sheet("Device Client API", client_data)
create_sheet("Admin Management API", admin_data)
create_sheet("Logging API", logging_data)

wb.save(filepath)
print("Updated Excel file with Device Info Sync API generated successfully!")

import os
import sys

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
    import openpyxl
import csv
from openpyxl.styles import Font, PatternFill, Alignment

base_dir = r"c:\Users\user\Documents\VSCode\CellbigCMS\CellbigCMS"
csv_path = os.path.join(base_dir, "api_specification.csv")
xlsx_path = os.path.join(base_dir, "api_specification.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Spec"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
alignment = Alignment(wrap_text=True, vertical="center")

with open(csv_path, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for i, row in enumerate(reader):
        ws.append(row)
        # Apply styles to the row
        for cell in ws[i+1]:
            cell.alignment = alignment
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill

# Adjust column widths
col_widths = {'A': 15, 'B': 10, 'C': 40, 'D': 35, 'E': 40, 'F': 40}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

wb.save(xlsx_path)
print("Excel file generated successfully!")
